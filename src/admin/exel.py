from openpyxl import load_workbook


def disc():
    wb = load_workbook('src/admin/Menu.xlsx')
    sheet = wb.active

    menus = []
    submenus = []
    dishes = []
    last_menu_id = None
    last_submenu_id = None

    for row in sheet.iter_rows(values_only=True):
        if row[0] is not None:  # Это меню
            last_menu_id = row[0]
            menus.append([row[0], row[1], row[2]])
        elif row[1] is not None:  # Это подменю
            last_submenu_id = row[1]
            submenus.append([last_menu_id, row[1], row[2], row[3]])
        else:  # Это блюдо
            dishes.append([last_menu_id, last_submenu_id, row[2], row[6]])

    return {'dishes': dishes}
